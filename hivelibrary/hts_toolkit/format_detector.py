from openpyxl import load_workbook






class HtsToolkit():
    supported_formats = [
        "BTK_BASIC_1",
        "BTK_STANDARD_1"
    ]
    
    not_support_returner = "UNSUPPORTED"
    
    def __init__(self):
        pass

    def detect_hts_record_formats(self,targetFilePath:str) -> str:
        # Check is excel file or other file[s]
        
        
        try:
            workbook = load_workbook(targetFilePath)
            workbook.close()
        except Exception as err:
            return self.not_support_returner
        
        workBook = load_workbook(targetFilePath)    
        workSheet = workbook.active
        
        try:
            # Ckeck for BTK_BASIC_1
            id_is = str(workSheet.cell(53,1).value)
            tarih = workSheet.cell(53,2).value
            imei = str(workSheet.cell(53,3).value)
            numara = str(workSheet.cell(53,4).value)
            control_basamak = workSheet.cell(53,5).value
            imei_proccessed = imei.replace("-","")
            imei_proccessed = imei_proccessed.replace(" ","")
            workBook.close()
            if id_is != None and tarih != None and imei != None and numara != None and control_basamak == None: 
                if id_is.isnumeric() and imei_proccessed.isnumeric() and len(imei_proccessed) == 15:
                    return "BTK_BASIC_1"
        
        except Exception as err:
            pass
        
        
        try:
            # Check for BTK_STANDARD_1
            id_is = str(workSheet.cell(52,1).value)
            source_number = str(workSheet.cell(52,2).value)
            islem_tipi = str(workSheet.cell(52,3).value)
            karsi_numara = str(workSheet.cell(52,4).value)
            islem_tarihi = str(workSheet.cell(52,5).value)
            islem_süresi = str(workSheet.cell(52,6).value)
            hedef_numara_ad_soyad = str(workSheet.cell(52,7).value)
            hedef_numara_tc_numarası = str(workSheet.cell(52,8).value)
            source_imei = str(workSheet.cell(52,9).value)
            islemGerceklesenBaz = str(workSheet.cell(52,10).value)
            
            control_basamak = workSheet.cell(52,11).value
            imei_proccessed = source_imei.replace("-","")
            imei_proccessed = imei_proccessed.replace(" ","")
            workBook.close()
            
            
            if id_is != None and islem_tarihi != None and source_imei != None and source_number != None and control_basamak == None and islem_tipi != None and islem_süresi != None and karsi_numara != None: 
                if id_is.isnumeric() and imei_proccessed.isnumeric() and len(imei_proccessed) == 15:
                    return "BTK_STANDARD_1"
            
        except Exception as err:
            pass
        
        
        
        return self.not_support_returner
            
if __name__ == "__main__":
    import sys
    targetFile = sys.argv[1]
    readerTool = HtsToolkit()
    print(readerTool.detect_hts_record_formats(targetFilePath=targetFile))